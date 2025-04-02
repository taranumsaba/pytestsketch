import openpyxl
import os.path


def get_cel_data(sheet, rowNo, colNo):
    p = os.getcwd().replace("\\Testcases", "")
    path1 = p.__add__("\\Config\\Users.xlsx")
    wb = openpyxl.load_workbook(path1)
    #rowNo = 2
    sh = wb[sheet]
    datalist = []
    col_count = sh.max_column

    row_count = sh.max_row
    if rowNo <= row_count and colNo <= col_count:
        for i in range(rowNo, rowNo + 1):
            for j in range(colNo, colNo + 1):
                datalist.append(sh.cell(row=i, column=j).value)
        print(datalist)
    else:
        print("Maximum Row reached")
    return [datalist]


def get_cel_data_rows():
    p = os.getcwd().replace("\\Testcases", "")
    path1 = p.__add__("\\Config\\Users.xlsx")
    wb = openpyxl.load_workbook(path1)
    #rowNo = 2
    #sh = wb[sheet]
    sh = wb.active

    mainlist = []
    blist = []
    col_count = sh.max_column
    print(col_count)
    row_count = sh.max_row
    print(row_count)
   
    for i in range(2, 3):
        datalist = []
        for j in range(1, col_count):
            datalist.append(sh.cell(row=i, column=j).value)
        mainlist.append(datalist)
        #print(mainlist)
    return mainlist











